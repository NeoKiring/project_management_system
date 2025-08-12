"""
Excelインポート機能
4フォーマット対応・自動検出・階層構造復元
"""

import os
import re
from datetime import datetime, date
from typing import Dict, List, Optional, Any, Tuple, Union
from pathlib import Path
import logging

try:
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    openpyxl = None
    load_workbook = None
    Worksheet = None

from ..core.logger import ProjectLogger, LogCategory
from ..core.error_handler import (
    handle_errors, validate_input, FileIOError, DataError, 
    ValidationError, RecoveryStrategy
)
from ..models.project import Project
from ..models.phase import Phase
from ..models.process import Process
from ..models.task import Task
from ..models.base import ProjectStatus, TaskStatus


class ExcelFormatType:
    """Excelフォーマット種別定義"""
    STANDARD = "standard"          # 標準フォーマット（スケジュール+入力シート）
    MSPROJECT = "msproject"        # MS Project類似（Tasks+Resources）
    SIMPLE = "simple"              # シンプルフォーマット（単一シート）
    CUSTOM = "custom"              # カスタムフォーマット（柔軟対応）


class ImportResult:
    """インポート結果クラス"""
    
    def __init__(self):
        self.success = False
        self.format_type = None
        self.imported_counts = {
            'projects': 0,
            'phases': 0,
            'processes': 0,
            'tasks': 0
        }
        self.errors = []
        self.warnings = []
        self.skipped_rows = []
        self.created_entities = {
            'projects': {},
            'phases': {},
            'processes': {},
            'tasks': {}
        }
        self.processing_time = 0.0
        self.file_info = {}
    
    def add_error(self, row_num: int, message: str, details: str = ""):
        """エラーを追加"""
        self.errors.append({
            'row': row_num,
            'message': message,
            'details': details,
            'timestamp': datetime.now().isoformat()
        })
    
    def add_warning(self, row_num: int, message: str):
        """警告を追加"""
        self.warnings.append({
            'row': row_num,
            'message': message,
            'timestamp': datetime.now().isoformat()
        })
    
    def to_dict(self) -> Dict[str, Any]:
        """辞書形式で取得"""
        return {
            'success': self.success,
            'format_type': self.format_type,
            'imported_counts': self.imported_counts.copy(),
            'error_count': len(self.errors),
            'warning_count': len(self.warnings),
            'errors': self.errors.copy(),
            'warnings': self.warnings.copy(),
            'skipped_rows': self.skipped_rows.copy(),
            'processing_time': self.processing_time,
            'file_info': self.file_info.copy()
        }


class ExcelFormatDetector:
    """Excelフォーマット自動検出エンジン"""
    
    def __init__(self, logger: ProjectLogger):
        self.logger = logger
        
        # フォーマット判定用のキーワード
        self.format_keywords = {
            ExcelFormatType.STANDARD: {
                'sheet_names': ['スケジュール', 'schedule', '入力', 'input'],
                'headers': ['プロジェクト', 'フェーズ', 'プロセス', 'タスク', '階層']
            },
            ExcelFormatType.MSPROJECT: {
                'sheet_names': ['tasks', 'resources', 'タスク', 'リソース'],
                'headers': ['task name', 'resource name', 'start', 'finish', 'duration']
            },
            ExcelFormatType.SIMPLE: {
                'sheet_names': ['data', 'sheet1', 'データ'],
                'headers': ['名前', 'name', '担当者', 'assignee', '期限', 'deadline']
            }
        }
    
    @handle_errors(recovery_strategy=RecoveryStrategy.FALLBACK, fallback_value=ExcelFormatType.CUSTOM)
    def detect_format(self, workbook: openpyxl.Workbook) -> str:
        """
        Excelファイルのフォーマットを自動検出
        
        Args:
            workbook: openpyxlワークブック
            
        Returns:
            検出されたフォーマット種別
        """
        sheet_names = [sheet.title.lower() for sheet in workbook.worksheets]
        self.logger.debug(
            LogCategory.DATA,
            f"シート名を検出: {sheet_names}",
            module="external.excel_importer"
        )
        
        format_scores = {}
        
        for format_type, keywords in self.format_keywords.items():
            score = 0
            
            # シート名による判定
            for keyword in keywords['sheet_names']:
                if any(keyword.lower() in sheet_name for sheet_name in sheet_names):
                    score += 2
            
            # ヘッダーによる判定
            for sheet in workbook.worksheets:
                headers = self._extract_headers(sheet)
                header_text = ' '.join(headers).lower()
                
                for keyword in keywords['headers']:
                    if keyword.lower() in header_text:
                        score += 1
            
            format_scores[format_type] = score
        
        # 最高スコアのフォーマットを選択
        detected_format = max(format_scores.items(), key=lambda x: x[1])
        
        self.logger.info(
            LogCategory.DATA,
            f"フォーマット検出結果: {detected_format[0]} (スコア: {detected_format[1]})",
            module="external.excel_importer",
            format_scores=format_scores
        )
        
        return detected_format[0] if detected_format[1] > 0 else ExcelFormatType.CUSTOM
    
    def _extract_headers(self, sheet: Worksheet, max_rows: int = 5) -> List[str]:
        """シートからヘッダー行を抽出"""
        headers = []
        
        for row_num in range(1, min(max_rows + 1, sheet.max_row + 1)):
            row_headers = []
            for col_num in range(1, min(20, sheet.max_column + 1)):  # 最大20列まで
                cell = sheet.cell(row=row_num, column=col_num)
                if cell.value:
                    row_headers.append(str(cell.value).strip())
            
            if len(row_headers) > len(headers):
                headers = row_headers
        
        return headers


class ExcelImporter:
    """
    Excel インポートメイン機能
    """
    
    def __init__(self, project_management_system=None):
        """
        インポーターの初期化
        
        Args:
            project_management_system: プロジェクト管理システムのインスタンス
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxlライブラリがインストールされていません。pip install openpyxl を実行してください。")
        
        self.pms = project_management_system
        self.logger = ProjectLogger()
        self.detector = ExcelFormatDetector(self.logger)
        
        # 設定値
        self.max_import_rows = 10000
        self.default_encoding = 'utf-8'
        self.date_formats = ['%Y-%m-%d', '%Y/%m/%d', '%m/%d/%Y', '%d/%m/%Y']
        
        # 処理統計
        self.import_stats = {
            'total_files': 0,
            'successful_imports': 0,
            'failed_imports': 0,
            'total_entities': 0
        }
        
        self.logger.info(
            LogCategory.SYSTEM,
            "Excelインポーター初期化完了",
            module="external.excel_importer"
        )
    
    @handle_errors(recovery_strategy=RecoveryStrategy.NONE)
    @validate_input()
    def import_from_file(self, file_path: str, format_type: str = None, 
                        options: Dict[str, Any] = None) -> ImportResult:
        """
        Excelファイルからデータをインポート
        
        Args:
            file_path: Excelファイルパス
            format_type: フォーマット種別（Noneの場合は自動検出）
            options: インポートオプション
            
        Returns:
            インポート結果
        """
        start_time = datetime.now()
        result = ImportResult()
        
        try:
            # ファイル存在チェック
            if not os.path.exists(file_path):
                raise FileIOError(f"ファイルが見つかりません: {file_path}", file_path=file_path)
            
            # ファイル情報取得
            file_info = Path(file_path)
            result.file_info = {
                'name': file_info.name,
                'size': file_info.stat().st_size,
                'modified': datetime.fromtimestamp(file_info.stat().st_mtime).isoformat()
            }
            
            self.logger.info(
                LogCategory.DATA,
                f"Excelインポート開始: {file_path}",
                module="external.excel_importer",
                file_info=result.file_info
            )
            
            # Excelファイル読み込み
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            
            # フォーマット検出
            if not format_type:
                format_type = self.detector.detect_format(workbook)
            
            result.format_type = format_type
            
            # フォーマット別インポート実行
            if format_type == ExcelFormatType.STANDARD:
                self._import_standard_format(workbook, result, options or {})
            elif format_type == ExcelFormatType.MSPROJECT:
                self._import_msproject_format(workbook, result, options or {})
            elif format_type == ExcelFormatType.SIMPLE:
                self._import_simple_format(workbook, result, options or {})
            else:  # CUSTOM
                self._import_custom_format(workbook, result, options or {})
            
            # ProjectManagementSystemに保存
            if self.pms and result.success:
                self._save_to_system(result)
            
            # 処理時間計算
            result.processing_time = (datetime.now() - start_time).total_seconds()
            
            # 統計更新
            self.import_stats['total_files'] += 1
            if result.success:
                self.import_stats['successful_imports'] += 1
                self.import_stats['total_entities'] += sum(result.imported_counts.values())
            else:
                self.import_stats['failed_imports'] += 1
            
            self.logger.info(
                LogCategory.DATA,
                f"Excelインポート完了: {result.success}, "
                f"entities={sum(result.imported_counts.values())}, "
                f"time={result.processing_time:.2f}s",
                module="external.excel_importer",
                result_summary=result.to_dict()
            )
            
        except Exception as e:
            result.success = False
            result.add_error(0, f"インポート処理エラー: {str(e)}")
            
            self.logger.error(
                LogCategory.ERROR,
                f"Excelインポートエラー: {e}",
                module="external.excel_importer",
                exception=e,
                file_path=file_path
            )
        
        return result
    
    def _import_standard_format(self, workbook: openpyxl.Workbook, 
                               result: ImportResult, options: Dict[str, Any]) -> None:
        """標準フォーマットのインポート処理"""
        
        # スケジュールシートを検索
        schedule_sheet = self._find_sheet(workbook, ['スケジュール', 'schedule', 'Schedule'])
        if not schedule_sheet:
            result.add_error(0, "スケジュールシートが見つかりません")
            return
        
        # ヘッダー行を検出
        header_row = self._detect_header_row(schedule_sheet)
        if not header_row:
            result.add_error(0, "ヘッダー行が見つかりません")
            return
        
        # 列マッピング検出
        column_mapping = self._detect_columns(schedule_sheet, header_row)
        
        # プロジェクト作成
        project_name = options.get('project_name', f"インポートプロジェクト_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
        project = Project(project_name, "Excelからインポートされたプロジェクト")
        result.created_entities['projects'][project.id] = project
        result.imported_counts['projects'] += 1
        
        # 階層データ読み込み
        current_phase = None
        current_process = None
        
        for row_num in range(header_row + 1, schedule_sheet.max_row + 1):
            try:
                row_data = self._extract_row_data(schedule_sheet, row_num, column_mapping)
                if not row_data or not row_data.get('name'):
                    continue
                
                # 階層レベルを判定
                level = self._detect_hierarchy_level(row_data)
                
                if level == 1:  # フェーズ
                    phase = Phase(row_data['name'], row_data.get('description', ''))
                    phase.parent_project_id = project.id
                    if row_data.get('end_date'):
                        phase.end_date = self._parse_date(row_data['end_date'])
                    
                    project.add_phase(phase.id)
                    result.created_entities['phases'][phase.id] = phase
                    result.imported_counts['phases'] += 1
                    current_phase = phase
                    current_process = None
                    
                elif level == 2 and current_phase:  # プロセス
                    process = Process(
                        row_data['name'], 
                        row_data.get('description', ''),
                        row_data.get('assignee', '担当者未設定')
                    )
                    process.parent_phase_id = current_phase.id
                    if row_data.get('start_date'):
                        process.start_date = self._parse_date(row_data['start_date'])
                    if row_data.get('end_date'):
                        process.end_date = self._parse_date(row_data['end_date'])
                    if row_data.get('estimated_hours'):
                        process.estimated_hours = self._parse_number(row_data['estimated_hours'])
                    
                    current_phase.add_process(process.id)
                    result.created_entities['processes'][process.id] = process
                    result.imported_counts['processes'] += 1
                    current_process = process
                    
                elif level == 3 and current_process:  # タスク
                    task = Task(row_data['name'], row_data.get('description', ''))
                    task.parent_process_id = current_process.id
                    if row_data.get('status'):
                        status_mapping = {
                            '未着手': TaskStatus.NOT_STARTED,
                            '進行中': TaskStatus.IN_PROGRESS,
                            '完了': TaskStatus.COMPLETED,
                            '対応不能': TaskStatus.CANNOT_HANDLE
                        }
                        task.status = status_mapping.get(row_data['status'], TaskStatus.NOT_STARTED)
                    
                    if row_data.get('priority'):
                        task.priority = int(row_data.get('priority', 3))
                    
                    current_process.add_task(task.id)
                    result.created_entities['tasks'][task.id] = task
                    result.imported_counts['tasks'] += 1
                
            except Exception as e:
                result.add_error(row_num, f"行処理エラー: {str(e)}")
                continue
        
        result.success = True
    
    def _import_msproject_format(self, workbook: openpyxl.Workbook, 
                                result: ImportResult, options: Dict[str, Any]) -> None:
        """MS Project類似フォーマットのインポート処理"""
        
        # Tasksシートを検索
        tasks_sheet = self._find_sheet(workbook, ['tasks', 'タスク', 'Tasks'])
        if not tasks_sheet:
            result.add_error(0, "Tasksシートが見つかりません")
            return
        
        # ヘッダー検出
        header_row = self._detect_header_row(tasks_sheet)
        column_mapping = self._detect_columns(tasks_sheet, header_row)
        
        # プロジェクト作成
        project_name = options.get('project_name', "MS Projectインポート")
        project = Project(project_name, "MS Project形式からインポート")
        
        # 単一フェーズ・プロセス作成（MS Project形式では階層が平坦なため）
        phase = Phase("メインフェーズ", "インポートされたメインフェーズ")
        phase.parent_project_id = project.id
        project.add_phase(phase.id)
        
        process = Process("メインプロセス", "インポートされたメインプロセス", "プロジェクトマネージャー")
        process.parent_phase_id = phase.id
        phase.add_process(process.id)
        
        # タスク読み込み
        for row_num in range(header_row + 1, tasks_sheet.max_row + 1):
            try:
                row_data = self._extract_row_data(tasks_sheet, row_num, column_mapping)
                if not row_data or not row_data.get('name'):
                    continue
                
                task = Task(row_data['name'], row_data.get('description', ''))
                task.parent_process_id = process.id
                
                # MS Project特有の属性をマッピング
                if row_data.get('duration'):
                    # 期間（日数）を工数（時間）に変換（1日=8時間と仮定）
                    duration_days = self._parse_number(row_data['duration'])
                    if duration_days:
                        task.estimated_hours = duration_days * 8
                
                if row_data.get('percent_complete'):
                    # 進捗率からステータスを推定
                    progress = self._parse_number(row_data['percent_complete'])
                    if progress == 100:
                        task.status = TaskStatus.COMPLETED
                    elif progress > 0:
                        task.status = TaskStatus.IN_PROGRESS
                
                process.add_task(task.id)
                result.created_entities['tasks'][task.id] = task
                result.imported_counts['tasks'] += 1
                
            except Exception as e:
                result.add_error(row_num, f"タスク処理エラー: {str(e)}")
        
        # エンティティ登録
        result.created_entities['projects'][project.id] = project
        result.created_entities['phases'][phase.id] = phase  
        result.created_entities['processes'][process.id] = process
        result.imported_counts['projects'] += 1
        result.imported_counts['phases'] += 1
        result.imported_counts['processes'] += 1
        
        result.success = True
    
    def _import_simple_format(self, workbook: openpyxl.Workbook, 
                             result: ImportResult, options: Dict[str, Any]) -> None:
        """シンプルフォーマットのインポート処理"""
        
        # 最初のシートを使用
        sheet = workbook.active
        header_row = self._detect_header_row(sheet)
        column_mapping = self._detect_columns(sheet, header_row)
        
        # プロジェクト作成
        project_name = options.get('project_name', "シンプルインポート")
        project = Project(project_name, "シンプル形式からインポート")
        
        # デフォルト階層作成
        phase = Phase("メインフェーズ", "")
        phase.parent_project_id = project.id
        project.add_phase(phase.id)
        
        # 各行をプロセスとして処理
        for row_num in range(header_row + 1, sheet.max_row + 1):
            try:
                row_data = self._extract_row_data(sheet, row_num, column_mapping)
                if not row_data or not row_data.get('name'):
                    continue
                
                process = Process(
                    row_data['name'],
                    row_data.get('description', ''),
                    row_data.get('assignee', '担当者未設定')
                )
                process.parent_phase_id = phase.id
                
                if row_data.get('deadline'):
                    process.end_date = self._parse_date(row_data['deadline'])
                
                phase.add_process(process.id)
                result.created_entities['processes'][process.id] = process
                result.imported_counts['processes'] += 1
                
            except Exception as e:
                result.add_error(row_num, f"プロセス処理エラー: {str(e)}")
        
        # エンティティ登録
        result.created_entities['projects'][project.id] = project
        result.created_entities['phases'][phase.id] = phase
        result.imported_counts['projects'] += 1
        result.imported_counts['phases'] += 1
        
        result.success = True
    
    def _import_custom_format(self, workbook: openpyxl.Workbook, 
                             result: ImportResult, options: Dict[str, Any]) -> None:
        """カスタムフォーマットのインポート処理"""
        
        # カスタムフォーマットでは最初のシートから可能な限りデータを抽出
        sheet = workbook.active
        header_row = self._detect_header_row(sheet)
        
        if not header_row:
            result.add_error(0, "ヘッダー行を検出できませんでした")
            return
        
        # 汎用的な列検出
        column_mapping = self._detect_columns_generic(sheet, header_row)
        
        # プロジェクト作成
        project_name = options.get('project_name', "カスタムインポート")
        project = Project(project_name, "カスタム形式からインポート")
        
        # 動的階層構築
        self._build_dynamic_hierarchy(sheet, header_row, column_mapping, project, result)
        
        result.success = True
    
    # ユーティリティメソッド
    def _find_sheet(self, workbook: openpyxl.Workbook, sheet_names: List[str]) -> Optional[Worksheet]:
        """指定された名前のシートを検索"""
        for sheet_name in sheet_names:
            for sheet in workbook.worksheets:
                if sheet_name.lower() in sheet.title.lower():
                    return sheet
        return None
    
    def _detect_header_row(self, sheet: Worksheet) -> Optional[int]:
        """ヘッダー行を検出"""
        for row_num in range(1, min(10, sheet.max_row + 1)):
            row_data = []
            for col_num in range(1, min(20, sheet.max_column + 1)):
                cell = sheet.cell(row=row_num, column=col_num)
                if cell.value:
                    row_data.append(str(cell.value).strip())
            
            # ヘッダーっぽい行を判定（複数の非空セルがある）
            if len(row_data) >= 3:
                return row_num
        
        return None
    
    def _detect_columns(self, sheet: Worksheet, header_row: int) -> Dict[str, int]:
        """列マッピングを検出"""
        column_mapping = {}
        
        # 標準的な列名パターン
        column_patterns = {
            'name': ['名前', 'name', 'タスク名', 'task', 'project', 'プロジェクト'],
            'description': ['説明', 'description', '備考', 'memo', 'note'],
            'assignee': ['担当者', 'assignee', 'resource', 'owner', '責任者'],
            'start_date': ['開始日', 'start', 'begin', '着手日'],
            'end_date': ['終了日', 'end', 'finish', '期限', 'deadline'],
            'status': ['状態', 'status', 'ステータス', '進捗状況'],
            'priority': ['優先度', 'priority', '重要度'],
            'estimated_hours': ['予想工数', 'estimated', '見積時間', 'duration'],
            'actual_hours': ['実績工数', 'actual', '実績時間'],
            'percent_complete': ['進捗率', 'progress', '完了率', '%']
        }
        
        for col_num in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=header_row, column=col_num)
            if not cell.value:
                continue
                
            header_text = str(cell.value).strip().lower()
            
            for field, patterns in column_patterns.items():
                for pattern in patterns:
                    if pattern.lower() in header_text:
                        column_mapping[field] = col_num
                        break
        
        return column_mapping
    
    def _detect_columns_generic(self, sheet: Worksheet, header_row: int) -> Dict[str, int]:
        """汎用的な列検出（カスタムフォーマット用）"""
        column_mapping = {}
        
        for col_num in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=header_row, column=col_num)
            if cell.value:
                header_text = str(cell.value).strip()
                column_mapping[f'col_{col_num}'] = col_num
                # 最初の列を名前として扱う
                if col_num == 1:
                    column_mapping['name'] = col_num
        
        return column_mapping
    
    def _extract_row_data(self, sheet: Worksheet, row_num: int, 
                         column_mapping: Dict[str, int]) -> Dict[str, Any]:
        """行データを抽出"""
        row_data = {}
        
        for field, col_num in column_mapping.items():
            cell = sheet.cell(row=row_num, column=col_num)
            if cell.value is not None:
                row_data[field] = str(cell.value).strip()
        
        return row_data
    
    def _detect_hierarchy_level(self, row_data: Dict[str, Any]) -> int:
        """階層レベルを検出"""
        name = row_data.get('name', '')
        
        # インデント数で判定
        indent_count = len(name) - len(name.lstrip())
        
        if indent_count == 0:
            return 1  # フェーズ
        elif indent_count <= 4:
            return 2  # プロセス
        else:
            return 3  # タスク
    
    def _parse_date(self, date_str: str) -> Optional[date]:
        """日付文字列をパース"""
        if not date_str:
            return None
            
        # Excelの日付セル値の場合
        if isinstance(date_str, datetime):
            return date_str.date()
        
        date_str = str(date_str).strip()
        
        for fmt in self.date_formats:
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue
        
        return None
    
    def _parse_number(self, number_str: str) -> Optional[float]:
        """数値文字列をパース"""
        if not number_str:
            return None
            
        try:
            # カンマ区切りを除去
            clean_str = str(number_str).replace(',', '').strip()
            return float(clean_str)
        except (ValueError, TypeError):
            return None
    
    def _build_dynamic_hierarchy(self, sheet: Worksheet, header_row: int, 
                                column_mapping: Dict[str, int], project: Project, 
                                result: ImportResult) -> None:
        """動的階層構築（カスタムフォーマット用）"""
        
        # デフォルト階層作成
        phase = Phase("インポートフェーズ", "カスタム形式からインポート")
        phase.parent_project_id = project.id
        project.add_phase(phase.id)
        
        process = Process("インポートプロセス", "", "システム")
        process.parent_phase_id = phase.id
        phase.add_process(process.id)
        
        # 各行をタスクとして処理
        for row_num in range(header_row + 1, sheet.max_row + 1):
            try:
                row_data = self._extract_row_data(sheet, row_num, column_mapping)
                if not row_data or not row_data.get('name'):
                    continue
                
                task = Task(row_data['name'], row_data.get('description', ''))
                task.parent_process_id = process.id
                process.add_task(task.id)
                
                result.created_entities['tasks'][task.id] = task
                result.imported_counts['tasks'] += 1
                
            except Exception as e:
                result.add_error(row_num, f"カスタム行処理エラー: {str(e)}")
        
        # エンティティ登録
        result.created_entities['projects'][project.id] = project
        result.created_entities['phases'][phase.id] = phase
        result.created_entities['processes'][process.id] = process
        result.imported_counts['projects'] += 1
        result.imported_counts['phases'] += 1
        result.imported_counts['processes'] += 1
    
    def _save_to_system(self, result: ImportResult) -> None:
        """ProjectManagementSystemにデータを保存"""
        try:
            if not self.pms:
                return
            
            # プロジェクトを保存
            for project in result.created_entities['projects'].values():
                self.pms.project_manager.projects[project.id] = project
                self.pms.data_store.save_project(project.id, project.to_dict())
            
            # フェーズを保存
            for phase in result.created_entities['phases'].values():
                self.pms.phase_manager.phases[phase.id] = phase
                self.pms.data_store.save_phase(phase.id, phase.to_dict())
            
            # プロセスを保存
            for process in result.created_entities['processes'].values():
                self.pms.process_manager.processes[process.id] = process
                self.pms.data_store.save_process(process.id, process.to_dict())
            
            # タスクを保存
            for task in result.created_entities['tasks'].values():
                self.pms.task_manager.tasks[task.id] = task
                self.pms.data_store.save_task(task.id, task.to_dict())
            
            self.logger.info(
                LogCategory.DATA,
                "インポートデータをシステムに保存完了",
                module="external.excel_importer",
                saved_counts=result.imported_counts
            )
            
        except Exception as e:
            self.logger.error(
                LogCategory.ERROR,
                f"システム保存エラー: {e}",
                module="external.excel_importer",
                exception=e
            )
            result.add_error(0, f"システム保存エラー: {str(e)}")
    
    def get_import_statistics(self) -> Dict[str, Any]:
        """インポート統計を取得"""
        return self.import_stats.copy()
    
    def __str__(self) -> str:
        """文字列表現"""
        return f"ExcelImporter(stats={self.import_stats})"