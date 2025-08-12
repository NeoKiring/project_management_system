"""
Excelエクスポート機能
4フォーマット対応・カスタマイズ・高品質出力
"""

import os
from datetime import datetime, date
from typing import Dict, List, Optional, Any, Tuple, Union
from pathlib import Path
import logging

try:
    import xlsxwriter
    XLSXWRITER_AVAILABLE = True
except ImportError:
    XLSXWRITER_AVAILABLE = False
    xlsxwriter = None

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
    from openpyxl.formatting.rule import ColorScaleRule
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    openpyxl = None

from ..core.logger import ProjectLogger, LogCategory
from ..core.error_handler import (
    handle_errors, validate_input, FileIOError, DataError, 
    ValidationError, RecoveryStrategy
)
from ..models.project import Project
from ..models.phase import Phase
from ..models.process import Process
from ..models.task import Task
from ..models.base import ProjectStatus, TaskStatus


class ExportFormat:
    """エクスポートフォーマット定義"""
    STANDARD = "standard"          # 標準フォーマット（スケジュール+入力シート）
    MSPROJECT = "msproject"        # MS Project類似（Tasks+Resources）
    SIMPLE = "simple"              # シンプルフォーマット（単一シート）
    CUSTOM = "custom"              # カスタムフォーマット（設定可能）


class ExportOptions:
    """エクスポートオプション"""
    
    def __init__(self):
        # 基本設定
        self.include_completed = True
        self.include_cancelled = False
        self.date_format = "%Y-%m-%d"
        self.encoding = "utf-8"
        
        # 表示設定
        self.show_progress_bars = True
        self.show_gantt_chart = False
        self.use_colors = True
        self.freeze_panes = True
        
        # フィルタ設定
        self.project_ids = []  # 空の場合は全プロジェクト
        self.status_filter = []
        self.assignee_filter = []
        self.date_range = None  # (start_date, end_date)
        
        # カスタム列設定
        self.custom_columns = []
        self.column_widths = {}
        
        # 出力設定
        self.sheet_names = {
            'schedule': 'スケジュール',
            'input': '入力データ',
            'tasks': 'タスク',
            'resources': 'リソース',
            'summary': 'サマリー'
        }
    
    def to_dict(self) -> Dict[str, Any]:
        """辞書形式で取得"""
        return {
            'include_completed': self.include_completed,
            'include_cancelled': self.include_cancelled,
            'date_format': self.date_format,
            'show_progress_bars': self.show_progress_bars,
            'show_gantt_chart': self.show_gantt_chart,
            'use_colors': self.use_colors,
            'freeze_panes': self.freeze_panes,
            'project_ids': self.project_ids.copy(),
            'status_filter': self.status_filter.copy(),
            'assignee_filter': self.assignee_filter.copy(),
            'custom_columns': self.custom_columns.copy(),
            'sheet_names': self.sheet_names.copy()
        }


class ExportResult:
    """エクスポート結果クラス"""
    
    def __init__(self):
        self.success = False
        self.file_path = ""
        self.format_type = ""
        self.exported_counts = {
            'projects': 0,
            'phases': 0,
            'processes': 0,
            'tasks': 0
        }
        self.file_size = 0
        self.processing_time = 0.0
        self.warnings = []
        self.errors = []
        
    def add_warning(self, message: str):
        """警告を追加"""
        self.warnings.append({
            'message': message,
            'timestamp': datetime.now().isoformat()
        })
    
    def add_error(self, message: str):
        """エラーを追加"""
        self.errors.append({
            'message': message,
            'timestamp': datetime.now().isoformat()
        })
    
    def to_dict(self) -> Dict[str, Any]:
        """辞書形式で取得"""
        return {
            'success': self.success,
            'file_path': self.file_path,
            'format_type': self.format_type,
            'exported_counts': self.exported_counts.copy(),
            'file_size': self.file_size,
            'processing_time': self.processing_time,
            'warning_count': len(self.warnings),
            'error_count': len(self.errors),
            'warnings': self.warnings.copy(),
            'errors': self.errors.copy()
        }


class ExcelStyleManager:
    """Excelスタイル管理クラス"""
    
    def __init__(self, use_xlsxwriter: bool = True):
        self.use_xlsxwriter = use_xlsxwriter and XLSXWRITER_AVAILABLE
        
        if self.use_xlsxwriter:
            self.styles = {}
        else:
            # openpyxlスタイル定義
            self.header_font = Font(bold=True, color="FFFFFF")
            self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            self.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # プログレスバー用の色
            self.progress_colors = {
                'high': PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),  # 赤
                'medium': PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),  # 黄
                'low': PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # 緑
            }
    
    def create_xlsxwriter_formats(self, workbook) -> Dict[str, Any]:
        """xlsxwriter用フォーマット作成"""
        if not self.use_xlsxwriter:
            return {}
        
        formats = {}
        
        # ヘッダー
        formats['header'] = workbook.add_format({
            'bold': True,
            'font_color': 'white',
            'bg_color': '#4472C4',
            'border': 1,
            'align': 'center',
            'valign': 'vcenter'
        })
        
        # 基本セル
        formats['cell'] = workbook.add_format({
            'border': 1,
            'valign': 'top'
        })
        
        # 日付
        formats['date'] = workbook.add_format({
            'border': 1,
            'num_format': 'yyyy-mm-dd'
        })
        
        # 数値
        formats['number'] = workbook.add_format({
            'border': 1,
            'num_format': '#,##0.0'
        })
        
        # パーセント
        formats['percent'] = workbook.add_format({
            'border': 1,
            'num_format': '0%'
        })
        
        # 階層レベル別
        formats['level1'] = workbook.add_format({
            'bold': True,
            'bg_color': '#E7E6E6',
            'border': 1,
            'indent': 0
        })
        
        formats['level2'] = workbook.add_format({
            'border': 1,
            'indent': 1
        })
        
        formats['level3'] = workbook.add_format({
            'border': 1,
            'indent': 2
        })
        
        # ステータス別色分け
        formats['status_completed'] = workbook.add_format({
            'border': 1,
            'bg_color': '#C6EFCE',  # 薄緑
            'font_color': '#006100'
        })
        
        formats['status_in_progress'] = workbook.add_format({
            'border': 1,
            'bg_color': '#FFEB9C',  # 薄黄
            'font_color': '#9C5700'
        })
        
        formats['status_not_started'] = workbook.add_format({
            'border': 1,
            'bg_color': '#FFC7CE',  # 薄赤
            'font_color': '#9C0006'
        })
        
        return formats


class ExcelExporter:
    """
    Excel エクスポートメイン機能
    """
    
    def __init__(self, project_management_system=None):
        """
        エクスポーターの初期化
        
        Args:
            project_management_system: プロジェクト管理システムのインスタンス
        """
        if not OPENPYXL_AVAILABLE and not XLSXWRITER_AVAILABLE:
            raise ImportError("openpyxl または xlsxwriter ライブラリが必要です。")
        
        self.pms = project_management_system
        self.logger = ProjectLogger()
        
        # xlsxwriterを優先使用
        self.use_xlsxwriter = XLSXWRITER_AVAILABLE
        self.style_manager = ExcelStyleManager(self.use_xlsxwriter)
        
        # エクスポート統計
        self.export_stats = {
            'total_exports': 0,
            'successful_exports': 0,
            'failed_exports': 0,
            'total_entities': 0
        }
        
        self.logger.info(
            LogCategory.SYSTEM,
            f"Excelエクスポーター初期化完了 (エンジン: {'xlsxwriter' if self.use_xlsxwriter else 'openpyxl'})",
            module="external.excel_exporter"
        )
    
    @handle_errors(recovery_strategy=RecoveryStrategy.NONE)
    @validate_input()
    def export_to_file(self, file_path: str, format_type: str = ExportFormat.STANDARD,
                      options: ExportOptions = None) -> ExportResult:
        """
        データをExcelファイルにエクスポート
        
        Args:
            file_path: 出力ファイルパス
            format_type: エクスポートフォーマット
            options: エクスポートオプション
            
        Returns:
            エクスポート結果
        """
        start_time = datetime.now()
        result = ExportResult()
        result.format_type = format_type
        
        if not options:
            options = ExportOptions()
        
        try:
            # データ取得・フィルタリング
            data = self._gather_data(options, result)
            if not data:
                result.add_error("エクスポート対象のデータがありません")
                return result
            
            self.logger.info(
                LogCategory.DATA,
                f"Excelエクスポート開始: {file_path} (フォーマット: {format_type})",
                module="external.excel_exporter",
                data_counts={k: len(v) for k, v in data.items()}
            )
            
            # フォーマット別エクスポート実行
            if format_type == ExportFormat.STANDARD:
                self._export_standard_format(file_path, data, options, result)
            elif format_type == ExportFormat.MSPROJECT:
                self._export_msproject_format(file_path, data, options, result)
            elif format_type == ExportFormat.SIMPLE:
                self._export_simple_format(file_path, data, options, result)
            elif format_type == ExportFormat.CUSTOM:
                self._export_custom_format(file_path, data, options, result)
            else:
                raise ValidationError(f"未対応のフォーマット: {format_type}")
            
            # ファイル情報取得
            if os.path.exists(file_path):
                result.file_path = file_path
                result.file_size = os.path.getsize(file_path)
                result.success = True
            
            # 処理時間計算
            result.processing_time = (datetime.now() - start_time).total_seconds()
            
            # 統計更新
            self.export_stats['total_exports'] += 1
            if result.success:
                self.export_stats['successful_exports'] += 1
                self.export_stats['total_entities'] += sum(result.exported_counts.values())
            else:
                self.export_stats['failed_exports'] += 1
            
            self.logger.info(
                LogCategory.DATA,
                f"Excelエクスポート完了: {result.success}, "
                f"entities={sum(result.exported_counts.values())}, "
                f"size={result.file_size:,} bytes, "
                f"time={result.processing_time:.2f}s",
                module="external.excel_exporter",
                result_summary=result.to_dict()
            )
            
        except Exception as e:
            result.success = False
            result.add_error(f"エクスポート処理エラー: {str(e)}")
            
            self.logger.error(
                LogCategory.ERROR,
                f"Excelエクスポートエラー: {e}",
                module="external.excel_exporter",
                exception=e,
                file_path=file_path
            )
        
        return result
    
    def _gather_data(self, options: ExportOptions, result: ExportResult) -> Dict[str, List]:
        """データを収集・フィルタリング"""
        if not self.pms:
            result.add_error("プロジェクト管理システムが設定されていません")
            return {}
        
        data = {
            'projects': [],
            'phases': [],
            'processes': [],
            'tasks': []
        }
        
        # プロジェクト取得
        all_projects = self.pms.get_all_projects()
        
        # プロジェクトフィルタ適用
        if options.project_ids:
            projects = [p for p in all_projects if p.id in options.project_ids]
        else:
            projects = all_projects
        
        for project in projects:
            # ステータスフィルタ
            if options.status_filter and project.status not in options.status_filter:
                continue
            
            # 完了・中止プロジェクトフィルタ
            if not options.include_completed and project.status == ProjectStatus.COMPLETED:
                continue
            if not options.include_cancelled and project.status == ProjectStatus.SUSPENDED:
                continue
            
            data['projects'].append(project)
            result.exported_counts['projects'] += 1
            
            # フェーズ取得
            for phase_id in project.phases:
                phase = self.pms.get_phase(phase_id)
                if phase:
                    data['phases'].append(phase)
                    result.exported_counts['phases'] += 1
                    
                    # プロセス取得
                    for process_id in phase.processes:
                        process = self.pms.get_process(process_id)
                        if process:
                            # 担当者フィルタ
                            if options.assignee_filter and process.assignee not in options.assignee_filter:
                                continue
                            
                            data['processes'].append(process)
                            result.exported_counts['processes'] += 1
                            
                            # タスク取得
                            for task_id in process.tasks:
                                task = self.pms.get_task(task_id)
                                if task:
                                    data['tasks'].append(task)
                                    result.exported_counts['tasks'] += 1
        
        return data
    
    def _export_standard_format(self, file_path: str, data: Dict[str, List], 
                               options: ExportOptions, result: ExportResult) -> None:
        """標準フォーマットエクスポート"""
        
        if self.use_xlsxwriter:
            self._export_standard_xlsxwriter(file_path, data, options, result)
        else:
            self._export_standard_openpyxl(file_path, data, options, result)
    
    def _export_standard_xlsxwriter(self, file_path: str, data: Dict[str, List], 
                                   options: ExportOptions, result: ExportResult) -> None:
        """xlsxwriterを使用した標準フォーマットエクスポート"""
        
        workbook = xlsxwriter.Workbook(file_path)
        formats = self.style_manager.create_xlsxwriter_formats(workbook)
        
        # スケジュールシート作成
        schedule_sheet = workbook.add_worksheet(options.sheet_names['schedule'])
        
        # ヘッダー行
        headers = [
            'レベル', '名前', '説明', '担当者', '開始日', '終了日',
            '進捗率', '状態', '予想工数', '実績工数', '優先度'
        ]
        
        for col, header in enumerate(headers):
            schedule_sheet.write(0, col, header, formats['header'])
        
        # 列幅設定
        column_widths = [8, 30, 40, 15, 12, 12, 10, 12, 12, 12, 8]
        for col, width in enumerate(column_widths):
            schedule_sheet.set_column(col, col, width)
        
        row = 1
        
        # データ出力
        for project in data['projects']:
            # プロジェクト行
            self._write_project_row_xlsxwriter(schedule_sheet, row, project, formats)
            row += 1
            
            # フェーズ出力
            for phase_id in project.phases:
                phase = next((p for p in data['phases'] if p.id == phase_id), None)
                if phase:
                    self._write_phase_row_xlsxwriter(schedule_sheet, row, phase, formats)
                    row += 1
                    
                    # プロセス出力
                    for process_id in phase.processes:
                        process = next((p for p in data['processes'] if p.id == process_id), None)
                        if process:
                            self._write_process_row_xlsxwriter(schedule_sheet, row, process, formats)
                            row += 1
                            
                            # タスク出力
                            for task_id in process.tasks:
                                task = next((t for t in data['tasks'] if t.id == task_id), None)
                                if task:
                                    self._write_task_row_xlsxwriter(schedule_sheet, row, task, formats)
                                    row += 1
        
        # フリーズペイン設定
        if options.freeze_panes:
            schedule_sheet.freeze_panes(1, 2)  # ヘッダー行と名前列を固定
        
        # 入力データシート作成
        self._create_input_sheet_xlsxwriter(workbook, data, options, formats)
        
        workbook.close()
    
    def _write_project_row_xlsxwriter(self, sheet, row: int, project: Project, formats: Dict) -> None:
        """プロジェクト行を出力（xlsxwriter）"""
        sheet.write(row, 0, 'P', formats['level1'])
        sheet.write(row, 1, project.name, formats['level1'])
        sheet.write(row, 2, project.description, formats['level1'])
        sheet.write(row, 3, project.manager, formats['level1'])
        sheet.write(row, 4, project.start_date.isoformat() if project.start_date else '', formats['date'])
        sheet.write(row, 5, project.end_date.isoformat() if project.end_date else '', formats['date'])
        sheet.write(row, 6, project.progress / 100, formats['percent'])
        sheet.write(row, 7, project.status, self._get_status_format(project.status, formats))
        sheet.write(row, 8, '', formats['cell'])  # 工数はプロジェクトレベルでは表示しない
        sheet.write(row, 9, '', formats['cell'])
        sheet.write(row, 10, project.priority, formats['cell'])
    
    def _write_phase_row_xlsxwriter(self, sheet, row: int, phase: Phase, formats: Dict) -> None:
        """フェーズ行を出力（xlsxwriter）"""
        sheet.write(row, 0, 'Ph', formats['level2'])
        sheet.write(row, 1, '  ' + phase.name, formats['level2'])
        sheet.write(row, 2, phase.description, formats['level2'])
        sheet.write(row, 3, '', formats['level2'])
        sheet.write(row, 4, '', formats['level2'])
        sheet.write(row, 5, phase.end_date.isoformat() if phase.end_date else '', formats['date'])
        sheet.write(row, 6, phase.progress / 100, formats['percent'])
        sheet.write(row, 7, phase.get_status(), self._get_status_format(phase.get_status(), formats))
        sheet.write(row, 8, '', formats['level2'])
        sheet.write(row, 9, '', formats['level2'])
        sheet.write(row, 10, phase.priority, formats['level2'])
    
    def _write_process_row_xlsxwriter(self, sheet, row: int, process: Process, formats: Dict) -> None:
        """プロセス行を出力（xlsxwriter）"""
        sheet.write(row, 0, 'Pr', formats['level3'])
        sheet.write(row, 1, '    ' + process.name, formats['level3'])
        sheet.write(row, 2, process.description, formats['level3'])
        sheet.write(row, 3, process.assignee, formats['level3'])
        sheet.write(row, 4, process.start_date.isoformat() if process.start_date else '', formats['date'])
        sheet.write(row, 5, process.end_date.isoformat() if process.end_date else '', formats['date'])
        sheet.write(row, 6, process.progress / 100, formats['percent'])
        sheet.write(row, 7, process.get_status(), self._get_status_format(process.get_status(), formats))
        sheet.write(row, 8, process.estimated_hours or '', formats['number'])
        sheet.write(row, 9, process.actual_hours or '', formats['number'])
        sheet.write(row, 10, process.priority, formats['level3'])
    
    def _write_task_row_xlsxwriter(self, sheet, row: int, task: Task, formats: Dict) -> None:
        """タスク行を出力（xlsxwriter）"""
        sheet.write(row, 0, 'T', formats['cell'])
        sheet.write(row, 1, '      ' + task.name, formats['cell'])
        sheet.write(row, 2, task.description, formats['cell'])
        sheet.write(row, 3, '', formats['cell'])
        sheet.write(row, 4, '', formats['cell'])
        sheet.write(row, 5, '', formats['cell'])
        sheet.write(row, 6, 1.0 if task.is_completed() else 0.0, formats['percent'])
        sheet.write(row, 7, task.status, self._get_status_format(task.status, formats))
        sheet.write(row, 8, task.estimated_hours or '', formats['number'])
        sheet.write(row, 9, task.actual_hours or '', formats['number'])
        sheet.write(row, 10, task.priority, formats['cell'])
    
    def _get_status_format(self, status: str, formats: Dict) -> Any:
        """ステータスに応じたフォーマットを取得"""
        if status in ['完了', TaskStatus.COMPLETED, '完成']:
            return formats.get('status_completed', formats['cell'])
        elif status in ['進行中', TaskStatus.IN_PROGRESS, '進行']:
            return formats.get('status_in_progress', formats['cell'])
        else:
            return formats.get('status_not_started', formats['cell'])
    
    def _create_input_sheet_xlsxwriter(self, workbook, data: Dict[str, List], 
                                      options: ExportOptions, formats: Dict) -> None:
        """入力データシートを作成（xlsxwriter）"""
        input_sheet = workbook.add_worksheet(options.sheet_names['input'])
        
        # ヘッダー行
        headers = [
            'ID', 'タイプ', '名前', '説明', '親ID', '担当者', 
            '開始日', '終了日', '予想工数', '実績工数', '進捗率', 
            '状態', '優先度', '作成日', '更新日'
        ]
        
        for col, header in enumerate(headers):
            input_sheet.write(0, col, header, formats['header'])
        
        # 列幅設定
        column_widths = [36, 8, 25, 40, 36, 15, 12, 12, 10, 10, 8, 12, 8, 16, 16]
        for col, width in enumerate(column_widths):
            input_sheet.set_column(col, col, width)
        
        row = 1
        
        # 全エンティティを出力
        all_entities = []
        
        # プロジェクト
        for project in data['projects']:
            all_entities.append(('Project', project, None))
        
        # フェーズ
        for phase in data['phases']:
            all_entities.append(('Phase', phase, phase.parent_project_id))
        
        # プロセス
        for process in data['processes']:
            all_entities.append(('Process', process, process.parent_phase_id))
        
        # タスク
        for task in data['tasks']:
            all_entities.append(('Task', task, task.parent_process_id))
        
        # データ出力
        for entity_type, entity, parent_id in all_entities:
            input_sheet.write(row, 0, entity.id, formats['cell'])
            input_sheet.write(row, 1, entity_type, formats['cell'])
            input_sheet.write(row, 2, entity.name, formats['cell'])
            input_sheet.write(row, 3, entity.description, formats['cell'])
            input_sheet.write(row, 4, parent_id or '', formats['cell'])
            
            # 担当者（プロセスのみ）
            assignee = getattr(entity, 'assignee', '') if entity_type == 'Process' else ''
            input_sheet.write(row, 5, assignee, formats['cell'])
            
            # 日付
            start_date = getattr(entity, 'start_date', None)
            end_date = getattr(entity, 'end_date', None)
            input_sheet.write(row, 6, start_date.isoformat() if start_date else '', formats['date'])
            input_sheet.write(row, 7, end_date.isoformat() if end_date else '', formats['date'])
            
            # 工数
            estimated_hours = getattr(entity, 'estimated_hours', None)
            actual_hours = getattr(entity, 'actual_hours', None)
            input_sheet.write(row, 8, estimated_hours or '', formats['number'])
            input_sheet.write(row, 9, actual_hours or '', formats['number'])
            
            # 進捗率
            progress = getattr(entity, 'progress', 0)
            input_sheet.write(row, 10, progress / 100, formats['percent'])
            
            # ステータス
            status = getattr(entity, 'status', '')
            input_sheet.write(row, 11, status, formats['cell'])
            
            # 優先度
            priority = getattr(entity, 'priority', '')
            input_sheet.write(row, 12, priority, formats['cell'])
            
            # 作成日・更新日
            input_sheet.write(row, 13, entity.created_at.isoformat(), formats['cell'])
            input_sheet.write(row, 14, entity.updated_at.isoformat(), formats['cell'])
            
            row += 1
        
        # フリーズペイン設定
        if options.freeze_panes:
            input_sheet.freeze_panes(1, 3)
    
    def _export_standard_openpyxl(self, file_path: str, data: Dict[str, List], 
                                 options: ExportOptions, result: ExportResult) -> None:
        """openpyxlを使用した標準フォーマットエクスポート"""
        # openpyxl実装（簡易版）
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = options.sheet_names['schedule']
        
        # ヘッダー行
        headers = ['レベル', '名前', '説明', '担当者', '開始日', '終了日', '進捗率', '状態']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = self.style_manager.header_font
            cell.fill = self.style_manager.header_fill
            cell.border = self.style_manager.border
        
        row = 2
        
        # データ出力（簡易版）
        for project in data['projects']:
            sheet.cell(row=row, column=1, value='P')
            sheet.cell(row=row, column=2, value=project.name)
            sheet.cell(row=row, column=3, value=project.description)
            sheet.cell(row=row, column=7, value=project.progress / 100)
            sheet.cell(row=row, column=8, value=project.status)
            row += 1
        
        workbook.save(file_path)
    
    def _export_msproject_format(self, file_path: str, data: Dict[str, List], 
                                options: ExportOptions, result: ExportResult) -> None:
        """MS Project類似フォーマットエクスポート"""
        
        if self.use_xlsxwriter:
            workbook = xlsxwriter.Workbook(file_path)
            formats = self.style_manager.create_xlsxwriter_formats(workbook)
            
            # Tasksシート
            tasks_sheet = workbook.add_worksheet(options.sheet_names['tasks'])
            
            # MS Project風ヘッダー
            headers = [
                'ID', 'Task Name', 'Duration', 'Start', 'Finish', 
                'Predecessors', 'Resource Names', '% Complete', 'Priority'
            ]
            
            for col, header in enumerate(headers):
                tasks_sheet.write(0, col, header, formats['header'])
            
            # 列幅設定
            column_widths = [8, 30, 12, 12, 12, 15, 20, 10, 8]
            for col, width in enumerate(column_widths):
                tasks_sheet.set_column(col, col, width)
            
            row = 1
            task_id = 1
            
            # フラットなタスクリストとして出力
            for project in data['projects']:
                for phase_id in project.phases:
                    phase = next((p for p in data['phases'] if p.id == phase_id), None)
                    if phase:
                        for process_id in phase.processes:
                            process = next((p for p in data['processes'] if p.id == process_id), None)
                            if process:
                                # プロセスをタスクとして出力
                                tasks_sheet.write(row, 0, task_id, formats['cell'])
                                tasks_sheet.write(row, 1, f"{project.name} - {phase.name} - {process.name}", formats['cell'])
                                
                                # 期間計算（開始日と終了日から）
                                if process.start_date and process.end_date:
                                    duration = (process.end_date - process.start_date).days + 1
                                else:
                                    duration = ''
                                
                                tasks_sheet.write(row, 2, duration, formats['number'])
                                tasks_sheet.write(row, 3, process.start_date.isoformat() if process.start_date else '', formats['date'])
                                tasks_sheet.write(row, 4, process.end_date.isoformat() if process.end_date else '', formats['date'])
                                tasks_sheet.write(row, 5, '', formats['cell'])  # Predecessors
                                tasks_sheet.write(row, 6, process.assignee, formats['cell'])
                                tasks_sheet.write(row, 7, process.progress / 100, formats['percent'])
                                tasks_sheet.write(row, 8, process.priority, formats['cell'])
                                
                                row += 1
                                task_id += 1
            
            # Resourcesシート
            resources_sheet = workbook.add_worksheet(options.sheet_names['resources'])
            
            resource_headers = ['Resource Name', 'Type', 'Max Units', 'Standard Rate', 'Overtime Rate']
            for col, header in enumerate(resource_headers):
                resources_sheet.write(0, col, header, formats['header'])
            
            # ユニークな担当者リストを作成
            assignees = set()
            for process in data['processes']:
                if process.assignee:
                    assignees.add(process.assignee)
            
            # リソース一覧出力
            for row, assignee in enumerate(sorted(assignees), 1):
                resources_sheet.write(row, 0, assignee, formats['cell'])
                resources_sheet.write(row, 1, 'Work', formats['cell'])
                resources_sheet.write(row, 2, '100%', formats['cell'])
                resources_sheet.write(row, 3, '¥0/h', formats['cell'])
                resources_sheet.write(row, 4, '¥0/h', formats['cell'])
            
            workbook.close()
        
        else:
            # openpyxl版の簡易実装
            self._export_msproject_openpyxl(file_path, data, options, result)
    
    def _export_msproject_openpyxl(self, file_path: str, data: Dict[str, List], 
                                  options: ExportOptions, result: ExportResult) -> None:
        """openpyxlでMS Project形式エクスポート（簡易版）"""
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = options.sheet_names['tasks']
        
        headers = ['ID', 'Task Name', 'Duration', 'Start', 'Finish', '% Complete']
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)
        
        row = 2
        task_id = 1
        
        for project in data['projects']:
            for process in data['processes']:
                sheet.cell(row=row, column=1, value=task_id)
                sheet.cell(row=row, column=2, value=process.name)
                sheet.cell(row=row, column=6, value=process.progress / 100)
                row += 1
                task_id += 1
        
        workbook.save(file_path)
    
    def _export_simple_format(self, file_path: str, data: Dict[str, List], 
                             options: ExportOptions, result: ExportResult) -> None:
        """シンプルフォーマットエクスポート"""
        
        if self.use_xlsxwriter:
            workbook = xlsxwriter.Workbook(file_path)
            formats = self.style_manager.create_xlsxwriter_formats(workbook)
            
            sheet = workbook.add_worksheet('データ')
            
            # シンプルなヘッダー
            headers = ['名前', '担当者', '期限', '状態', '進捗率', '備考']
            
            for col, header in enumerate(headers):
                sheet.write(0, col, header, formats['header'])
            
            row = 1
            
            # プロセスとタスクを平坦に出力
            for process in data['processes']:
                sheet.write(row, 0, process.name, formats['cell'])
                sheet.write(row, 1, process.assignee, formats['cell'])
                sheet.write(row, 2, process.end_date.isoformat() if process.end_date else '', formats['date'])
                sheet.write(row, 3, process.get_status(), formats['cell'])
                sheet.write(row, 4, process.progress / 100, formats['percent'])
                sheet.write(row, 5, process.description, formats['cell'])
                row += 1
            
            for task in data['tasks']:
                sheet.write(row, 0, f"  {task.name}", formats['cell'])
                sheet.write(row, 1, '', formats['cell'])
                sheet.write(row, 2, '', formats['cell'])
                sheet.write(row, 3, task.status, formats['cell'])
                sheet.write(row, 4, 1.0 if task.is_completed() else 0.0, formats['percent'])
                sheet.write(row, 5, task.description, formats['cell'])
                row += 1
            
            workbook.close()
        
        else:
            # openpyxl版（簡易実装）
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = 'データ'
            
            headers = ['名前', '担当者', '期限', '状態', '進捗率']
            for col, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col, value=header)
            
            row = 2
            for process in data['processes']:
                sheet.cell(row=row, column=1, value=process.name)
                sheet.cell(row=row, column=2, value=process.assignee)
                sheet.cell(row=row, column=4, value=process.get_status())
                sheet.cell(row=row, column=5, value=process.progress / 100)
                row += 1
            
            workbook.save(file_path)
    
    def _export_custom_format(self, file_path: str, data: Dict[str, List], 
                             options: ExportOptions, result: ExportResult) -> None:
        """カスタムフォーマットエクスポート"""
        
        # カスタム列が設定されていない場合はデフォルト列を使用
        if not options.custom_columns:
            options.custom_columns = [
                'name', 'description', 'assignee', 'start_date', 'end_date',
                'progress', 'status', 'priority', 'estimated_hours', 'actual_hours'
            ]
        
        if self.use_xlsxwriter:
            workbook = xlsxwriter.Workbook(file_path)
            formats = self.style_manager.create_xlsxwriter_formats(workbook)
            
            sheet = workbook.add_worksheet('カスタムデータ')
            
            # カスタムヘッダー
            for col, column_name in enumerate(options.custom_columns):
                header_text = self._get_column_display_name(column_name)
                sheet.write(0, col, header_text, formats['header'])
                
                # 列幅設定
                width = options.column_widths.get(column_name, 15)
                sheet.set_column(col, col, width)
            
            row = 1
            
            # 全エンティティを出力
            all_entities = (
                [('Project', p) for p in data['projects']] +
                [('Phase', p) for p in data['phases']] +
                [('Process', p) for p in data['processes']] +
                [('Task', t) for t in data['tasks']]
            )
            
            for entity_type, entity in all_entities:
                for col, column_name in enumerate(options.custom_columns):
                    value = self._get_entity_value(entity, column_name, entity_type)
                    
                    # データ型に応じたフォーマット適用
                    if column_name in ['start_date', 'end_date', 'created_at', 'updated_at']:
                        sheet.write(row, col, value, formats['date'])
                    elif column_name in ['progress']:
                        numeric_value = float(value) / 100 if value and str(value).replace('.', '').isdigit() else 0
                        sheet.write(row, col, numeric_value, formats['percent'])
                    elif column_name in ['estimated_hours', 'actual_hours', 'priority']:
                        sheet.write(row, col, value or '', formats['number'])
                    else:
                        sheet.write(row, col, value or '', formats['cell'])
                
                row += 1
            
            workbook.close()
        
        else:
            # openpyxl版カスタムエクスポート
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = 'カスタムデータ'
            
            # ヘッダー行
            for col, column_name in enumerate(options.custom_columns, 1):
                header_text = self._get_column_display_name(column_name)
                sheet.cell(row=1, column=col, value=header_text)
            
            # データ行
            row = 2
            all_entities = (
                [('Project', p) for p in data['projects']] +
                [('Phase', p) for p in data['phases']] +
                [('Process', p) for p in data['processes']] +
                [('Task', t) for t in data['tasks']]
            )
            
            for entity_type, entity in all_entities:
                for col, column_name in enumerate(options.custom_columns, 1):
                    value = self._get_entity_value(entity, column_name, entity_type)
                    sheet.cell(row=row, column=col, value=value)
                row += 1
            
            workbook.save(file_path)
    
    def _get_column_display_name(self, column_name: str) -> str:
        """列名の表示名を取得"""
        display_names = {
            'name': '名前',
            'description': '説明',
            'assignee': '担当者',
            'start_date': '開始日',
            'end_date': '終了日',
            'progress': '進捗率',
            'status': 'ステータス',
            'priority': '優先度',
            'estimated_hours': '予想工数',
            'actual_hours': '実績工数',
            'created_at': '作成日',
            'updated_at': '更新日',
            'entity_type': 'タイプ',
            'parent_id': '親ID'
        }
        return display_names.get(column_name, column_name)
    
    def _get_entity_value(self, entity, column_name: str, entity_type: str) -> Any:
        """エンティティから指定列の値を取得"""
        if column_name == 'entity_type':
            return entity_type
        elif column_name == 'parent_id':
            return getattr(entity, 'parent_project_id', None) or \
                   getattr(entity, 'parent_phase_id', None) or \
                   getattr(entity, 'parent_process_id', None) or ''
        elif hasattr(entity, column_name):
            value = getattr(entity, column_name)
            
            # 日付処理
            if isinstance(value, (date, datetime)):
                return value.isoformat()
            
            return value
        else:
            return ''
    
    def get_export_statistics(self) -> Dict[str, Any]:
        """エクスポート統計を取得"""
        return self.export_stats.copy()
    
    def __str__(self) -> str:
        """文字列表現"""
        return f"ExcelExporter(engine={'xlsxwriter' if self.use_xlsxwriter else 'openpyxl'}, stats={self.export_stats})"